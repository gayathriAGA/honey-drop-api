from datetime import datetime
from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from django.contrib.auth import authenticate
from django_filters.rest_framework import DjangoFilterBackend

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.decorators import api_view, permission_classes


from .permissions import ManageProducts, ManageLeads, ManageUsers, ManageCategories, ManageCustomers
from .models import User, Category, SubCategory, Product, Lead, Customer, ProductInterests, CustomerProducts
from .serializers import (
    UserSerializer, CategorySerializer, SubCategorySerializer, LoginSerializer,
    ProductSerializer, LeadSerializer, CustomerSerializer, ConvertLeadSerializer
)


@api_view(['GET'])
@permission_classes([AllowAny])
def health(request):
    return Response("OK", status=status.HTTP_200_OK)


class AuthViewSet(viewsets.ViewSet):
    """
    Auth ViewSet providing:
    - POST /api/auth/login/
    - POST /api/auth/logout/
    - GET /api/auth/me/
    """

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def login(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        user = authenticate(username=username, password=password)
        if not user:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        if user.role != "admin" and not user.customer_profile.is_active:
            return Response({
                "error": "Your account is blocked currently. Please contact your administrator."
            }, status=status.HTTP_401_UNAUTHORIZED)

        refresh = RefreshToken.for_user(user)
        data = UserSerializer(user).data
        data['token'] = str(refresh.access_token)
        # 'refresh': str(refresh),
        resp = Response(data, status=status.HTTP_200_OK)
        resp.set_cookie("token", refresh.access_token)
        return resp

    @action(detail=False, methods=['post'])
    def logout(self, request):
        try:
            # refresh_token = request.data.get("refresh")
            # token = RefreshToken(refresh_token)
            # token.blacklist()  # Only works if you enable token blacklisting
            return Response({'message': 'Logged out successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': 'Invalid token'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def me(self, request):
        user = request.user
        serializer = UserSerializer(user)
        return Response(serializer.data)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'status']
    search_fields = ['name', 'email']
    ordering_fields = ['created_at', 'name']
    permission_classes = [ManageUsers]
    # pagination_class = PageNumberPagination


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    permission_classes = [ManageCategories]


class SubCategoryViewSet(viewsets.ModelViewSet):
    queryset = SubCategory.objects.all()
    serializer_class = SubCategorySerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['category', 'status']
    search_fields = ['name', 'description']
    ordering_fields = ['name']
    permission_classes = [ManageCategories]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get('categoryId')
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        return queryset


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['name', 'specifications']
    ordering_fields = ['name', 'price', 'created_at']
    permission_classes = [ManageProducts]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get('categoryId')
        subcategory = self.request.query_params.get('subCategoryId')
        
        if category:
            queryset = queryset.filter(sub_category__category=category)
        if subcategory:
            queryset = queryset.filter(sub_category=subcategory)
        
        return queryset


class LeadViewSet(viewsets.ModelViewSet):
    queryset = Lead.objects.all().prefetch_related('interests__product')
    serializer_class = LeadSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'area', 'priority', 'source', 'follow_up_date']
    search_fields = ['name', 'phone', 'email', 'product', 'notes']
    ordering_fields = ['created_at', 'follow_up_date']
    permission_classes = [ManageLeads]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        sales_rep = self.request.query_params.get('salesRep')
        from_date = self.request.query_params.get('fromDate')
        to_date = self.request.query_params.get('toDate')
        
        if sales_rep:
            queryset = queryset.filter(sales_rep=sales_rep)
        if from_date:
            queryset = queryset.filter(created_at__gte=from_date)
        if to_date:
            queryset = queryset.filter(created_at__lte=to_date)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def convert(self, request, pk=None):
        lead = self.get_object()
        serializer = ConvertLeadSerializer(data=request.data)
        
        if serializer.is_valid():
            installation_date = serializer.validated_data['installationDate']
            warranty_years = serializer.validated_data['warrantyYears']
            expiry_date = installation_date + relativedelta(years=warranty_years)
            
            customer = Customer.objects.create(
                name=lead.name,
                phone=lead.phone,
                email=lead.email,
                area=lead.area,
                address=lead.address,
                installation_date=installation_date,
                expiry_date=expiry_date,
                amount=0,
                status='active',
                sales_rep=lead.sales_rep,
                notes=lead.notes
            )

            for product_interest in ProductInterests.objects.filter(lead=lead).all():
                CustomerProducts.objects.create(
                    customer=customer,
                    product=product_interest.product,
                )
            
            lead.status = 'won'
            lead.save()
            
            return Response(
                CustomerSerializer(customer).data,
                status=status.HTTP_201_CREATED
            )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def upload(self, request):
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        imported = 0
        failed = 0
        errors = []
        
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = dict(zip(headers, row))
                    
                    lead_data = {
                        'name': data.get('name'),
                        'phone': data.get('phone'),
                        'email': data.get('email'),
                        'area': data.get('area'),
                        'address': data.get('address'),
                        'status': data.get('status', 'new'),
                        'source': data.get('source'),
                        'priority': data.get('priority', 'medium'),
                        'notes': data.get('notes'),
                        'sales_rep': data.get('salesRep'),
                    }
                    
                    if data.get('followUpDate'):
                        lead_data['follow_up_date'] = data['followUpDate']

                    product_interests = []
                    for product_name in data.get('products').split(','):
                        product = Product.objects.filter(name=product_name).first()
                        if product is None:
                            raise Exception(f"Product `{product_name}` not found")
                        product_interests.append(product)

                    lead = Lead.objects.create(**lead_data)
                    for product_interest in product_interests:
                        ProductInterests.objects.create(
                            lead=lead,
                            product=product_interest,
                        )
                    imported += 1
                    
                except Exception as e:
                    failed += 1
                    errors.append({'row': idx, 'error': str(e)})
            
            return Response({
                'success': True,
                'imported': imported,
                'failed': failed,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'area']
    search_fields = ['name', 'phone', 'email', 'product', 'notes']
    ordering_fields = ['created_at', 'installation_date', 'expiry_date']
    permission_classes = [ManageCustomers]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        sales_rep = self.request.query_params.get('salesRep')
        
        if sales_rep:
            queryset = queryset.filter(sales_rep=sales_rep)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def upload(self, request):
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        imported = 0
        failed = 0
        errors = []
        
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = dict(zip(headers, row))
                    
                    installation_date = data.get('installationDate')
                    warranty_years = data.get('warrantyYears', 2)
                    
                    if isinstance(installation_date, str):
                        installation_date = datetime.strptime(installation_date, '%Y-%m-%d').date()
                    
                    expiry_date = installation_date + relativedelta(years=warranty_years)
                    
                    customer_data = {
                        'name': data.get('name'),
                        'phone': data.get('phone'),
                        'email': data.get('email'),
                        'area': data.get('area'),
                        'address': data.get('address'),
                        'installation_date': installation_date,
                        'expiry_date': expiry_date,
                        'amount': data.get('amount', 0),
                        'status': data.get('status', 'active'),
                        'sales_rep': data.get('salesRep'),
                        'notes': data.get('notes'),
                    }

                    product_interests = []
                    for product_name in data.get('products', '').split(","):
                        product = Product.objects.filter(name=product_name).first()
                        if product is None:
                            raise Exception(f"Product `{product_name}` not found")
                        product_interests.append(product)

                    customer = Customer.objects.create(**customer_data)
                    for product_interest in product_interests:
                        CustomerProducts.objects.create(
                            customer=customer,
                            product=product_interest,
                        )

                    imported += 1
                    
                except Exception as e:
                    failed += 1
                    errors.append({'row': idx, 'error': str(e)})
            
            return Response({
                'success': True,
                'imported': imported,
                'failed': failed,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
